#!/usr/bin/env python3
"""성적표 엑셀(.xlsm)에서 **익명 통계**만 뽑아 `cohort/baseline.json` 을 만든다.

왜 필요한가
-----------
석차·백분위·또래 정답률은 그때까지 이 브라우저에 채점해 둔 학생만 모집단으로
삼았다. 그래서 같은 시험인데 학생마다 인원이 달랐다 — 먼저 채점한 학생은
`3/9`, 나중에 채점한 학생은 `8/10`. 정답률도 마찬가지로 서너 명이 만든
숫자였다. 모집단이 "지금까지 내가 채점한 만큼"이면 그 숫자들은 뜻을 잃는다.

시험지 한 회차를 실제로 응시한 사람은 성적표 엑셀에 다 들어 있다. 그것을
모집단으로 쓰면 누구를 먼저 채점하든 같은 숫자가 나온다.

무엇을 담고, 무엇을 안 담는가
-----------------------------
**이름·학교·학년·수험번호는 담지 않는다.** 이 저장소는 공개되어 있다.
담는 것은 회차별 집계뿐이다.

    n     그 회차 응시 인원
    hist  맞은 문항 수 → 사람 수          (석차·백분위)
    qc    문항별 정답자 수                 (또래 정답률)
    q     문항별 [①,②,③,④,무응답] 사람 수  (선택 분포)

개인은 복원되지 않는다 — 누가 몇 번을 골랐는지는 어디에도 없다.
엑셀 원본은 저장소에 넣지 않는다.

총점은 180점 만점(문항당 3점)이라 3으로 나누면 맞은 문항 수가 된다.
3의 배수가 아닌 값이 있으면 감점이 섞였다는 뜻이므로 세지 않고 알린다.

`q` 의 합이 `n` 보다 작을 수 있다
---------------------------------
엑셀은 '모두정답'·'전원정답'·'1또는2' 처리된 문항의 **학생 답을 그 문자열로
덮어썼다.** 누가 몇 번을 골랐는지가 원본에서 이미 사라졌다는 뜻이다.

덮인 **칸만** 빼고 나머지는 그대로 센다. 처음에는 그런 칸이 하나라도 있으면
문항 전체를 버렸는데, 7회 52번은 42명 중 딱 한 칸이 덮였을 뿐인데 멀쩡한
41명을 함께 버리고 있었다. 그래서 `q` 의 합이 `n` 보다 작을 수 있고,
그 문항의 선택 분포 분모는 `n` 이 아니라 **`q` 의 합**이다.

한 칸도 안 남은 문항(전원정답처럼 전부 덮인 경우)만 `null` 이다. 그런
문항은 어차피 모든 보기가 정답이라 분포를 따질 것이 없다.
정답자 수(`qc`)는 채점 열에 남아 있으므로 그 문항도 또래 정답률에는 들어간다.

정답 여부는 그 회차의 **채점 열**을 그대로 쓴다. 앱의 채점 규칙으로 다시
세지 않는다 — 2020년에 실제로 매겨진 점수가 그 학생들의 사실이고, 답안이
덮인 문항은 다시 셀 수도 없다. 대신 답이 온전한 문항에서는 두 방식이 같은
답을 내는지 확인하고, 어긋나면 알린다(정답 키가 어긋났다는 뜻이다).

사용:
    python3 tools/gen_cohort_baseline.py <엑셀들이 있는 디렉터리> [--write]

    같은 회차 파일이 여럿이면(11회 0804 · 0811 처럼) 사람이 많은 쪽을 쓴다.
    나중에 받은 파일이 대개 몇 명 더 들어와 있다.

    --exam <회차id>   파일 이름에서 회차를 못 읽을 때 사람이 알려 준다.
                      (이름 규칙은 JMChC 것뿐이라 나머지 시험은 이게 필요하다)
    --merge           있던 기준 기록을 남기고 **이번에 읽은 회차만** 얹는다.
                      엑셀이 회차마다 따로 오므로 이것이 보통의 쓰임이다.

    보기:
        python3 tools/gen_cohort_baseline.py ~/엑셀 --exam sanyeom-60 --merge --write
"""

from __future__ import annotations

import collections
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "cohort" / "baseline.json"

SHEET = "성적입력"
NAME_COL, TOTAL_COL = 1, 6      # 0-based: 성명 · 총점
FIRST_ROW = 7                   # 6행이 머리글
PER_Q = 3                       # 문항당 배점


class Round:
    """한 회차 한 파일에서 뽑은 것."""

    def __init__(self, path: Path):
        self.path = path
        self.scores: list[int] = []          # 맞은 문항 수
        self.nq = 0
        self.qc: list[int] = []              # 문항별 정답자 수
        self.qopt: list[list[int] | None] = []   # 문항별 [①②③④,무응답] · 복원 불가면 None
        self.key: list[int | str] = []
        self.dirty: dict[int, int] = {}      # 문항 → 답이 덮인 사람 수


def _header_row(grid: list[list]) -> int:
    for r, line in enumerate(grid):
        if len(line) > 1 and line[1] == "성명":
            return r
    raise LookupError("'성명' 머리글을 못 찾았다")


def _block(header: list, label: str, nq_hint: int | None = None) -> tuple[int, int]:
    """머리글에서 `label` 뒤에 1,2,3… 이 이어지는 자리를 찾아 (시작 열, 문항 수)."""
    for c, v in enumerate(header):
        if v != label:
            continue
        n = 0
        while c + 1 + n < len(header) and header[c + 1 + n] == n + 1:
            n += 1
        if n >= 10 and (nq_hint is None or n == nq_hint):
            return c + 1, n
    raise LookupError(f"'{label}' 문항 블록을 못 찾았다")


def read_round(path: Path) -> Round | None:
    import openpyxl

    book = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in book.sheetnames:
        print(f"  건너뜀('{SHEET}' 시트 없음) {path.name}")
        return None
    grid = [list(r) for r in book[SHEET].iter_rows(values_only=True)]
    hr = _header_row(grid)
    header = grid[hr]
    acol, nq = _block(header, "번호")
    scol, _ = _block(header, "채점", nq)
    # ── 맞은 문항 수는 **직접 센다** ────────────────────────────────
    # 예전에는 '총점' 을 3으로 나눴다(문항당 3점, 180점 만점). 그런데 감점이
    # 있는 서식에서는 총점이 3의 배수가 아니고 음수도 나온다 — 산과염기 60제가
    # 그랬고, 23명 중 21명이 "총점이 이상하다" 며 통째로 버려졌다.
    #
    # 감점은 그 시험지의 채점 규칙이지 **맞은 개수**가 아니다. 석차·백분위가
    # 쓰는 것은 맞은 개수다. 그러니 감점을 거치지 말고, 정답 행과 학생 답을
    # 맞대어 그 자리에서 센다. 서식이 어떻든 같은 답이 나온다.
    #
    # 정답 행에 '모두정답'·'1또는2' 처럼 여러 답이 적힌 칸도 있다 — 그 칸에
    # 든 숫자를 모두 정답으로 본다. 숫자가 하나도 없으면(전원정답) 누구나
    # 맞은 것으로 센다.
    ans_key = []
    for i in range(nq):
        raw = grid[hr - 2][acol + i] if acol + i < len(grid[hr - 2]) else None
        if isinstance(raw, (int, float)):
            ans_key.append({int(raw)})
        else:
            nums = {int(x) for x in re.findall(r"[1-4]", str(raw or ""))}
            ans_key.append(nums)          # 빈 집합 = 전원정답
    # 서식에 '맞은문항수' 열이 있으면 우리가 센 것과 맞는지만 견준다.
    ccol = None
    for c, v in enumerate(header):
        if isinstance(v, str) and v.replace(" ", "") == "맞은문항수":
            ccol = c
            break
    mismatched = 0

    out = Round(path)
    out.nq = nq
    out.key = [grid[hr - 2][acol + i] for i in range(nq)]
    opt = [[0, 0, 0, 0, 0] for _ in range(nq)]      # ①②③④ + 무응답
    correct = [0] * nq
    dirty: collections.Counter = collections.Counter()

    for line in grid[hr + 1:]:
        if len(line) <= TOTAL_COL or not line[NAME_COL]:
            continue
        mine = 0
        for i in range(nq):
            a = line[acol + i] if acol + i < len(line) else None
            ok = False
            if not ans_key[i]:
                ok = True                            # 전원정답
            elif isinstance(a, (int, float)) and int(a) in ans_key[i]:
                ok = True
            if ok:
                mine += 1
                correct[i] += 1
            if a is None or a == 0:
                opt[i][4] += 1
            elif isinstance(a, int) and 1 <= a <= 4:
                opt[i][a - 1] += 1
            else:
                dirty[i] += 1                       # '모두정답' 등으로 덮인 칸
        out.scores.append(mine)
        if ccol is not None and ccol < len(line):
            said = line[ccol]
            if isinstance(said, (int, float)) and int(said) != mine:
                mismatched += 1

    if not out.scores:
        return None
    if mismatched:
        # 엑셀이 적어 둔 맞은 개수와 우리가 센 것이 다르다. 정답 행이 바뀌었거나
        # 그 서식이 다른 규칙으로 세고 있다는 뜻이라, 조용히 넘기면 안 된다.
        print(f"  ! 엑셀의 '맞은문항수' 와 다른 학생 {mismatched}명: {path.name}")
    out.qc = correct
    # 덮인 칸만 빠진다. 한 칸도 안 남았으면(전원정답 등) 그 문항만 null.
    out.qopt = [None if sum(opt[i]) == 0 else opt[i] for i in range(nq)]
    out.dirty = {i + 1: c for i, c in sorted(dirty.items())}
    return out


def cross_check(exam_id: str, rnd: Round) -> list[str]:
    """답이 온전한 문항에서 엑셀 채점과 앱의 정답 규칙이 같은 말을 하는지 본다."""
    path = ROOT / "exams.json"
    exams = {e["id"]: e for e in json.loads(path.read_text(encoding="utf-8"))}
    exam = exams.get(exam_id)
    if not exam:
        return [f"{exam_id}: exams.json 에 없다"]
    if exam["nQ"] != rnd.nq:
        return [f"{exam_id}: 문항 수가 다르다 (엑셀 {rnd.nq} · 앱 {exam['nQ']})"]

    multi = exam.get("multi") or {}
    bad = []
    for i in range(rnd.nq):
        if rnd.qopt[i] is None or rnd.dirty.get(i + 1):
            continue                                  # 답이 덮인 칸이 있으면 비교 못 한다
        acc = set(multi.get(str(i + 1)) or [int(exam["key"][i])])
        mine = sum(rnd.qopt[i][k] for k in range(4) if k + 1 in acc)
        if mine != rnd.qc[i]:
            bad.append(f"{exam_id} {i+1}번: 엑셀 채점 {rnd.qc[i]}명 · 앱 규칙 {mine}명")
    return bad


def build(folder: Path, force_id=None) -> dict:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("openpyxl 이 필요하다:  pip install openpyxl", file=sys.stderr)
        raise SystemExit(2)

    best: dict[str, Round] = {}
    for path in sorted(folder.rglob("*.xls*")):
        # 파일 이름에서 회차를 읽는 규칙은 JMChC 것뿐이다. 다른 시험(산과염기
        # 60제 …)은 이름 규칙이 없어 통째로 건너뛰었고, 그 회차는 기준 기록에
        # 아예 안 들어갔다 — 석차 모집단이 '앱으로 채점한 몇 명'만 남는다.
        # 실제로 산과염기 60제가 그래서 1/1 로 나왔다.
        # 이름으로 못 읽으면 --exam 으로 사람이 알려 준다.
        if force_id:
            exam_id = force_id
        else:
            found = re.search(r"JMchC\s+([\d-]+)\s*[#사]", path.name, re.IGNORECASE)
            if not found:
                print(f"  건너뜀(회차를 못 읽음) {path.name}"
                      "   ← --exam <회차id> 로 알려 줄 수 있다")
                continue
            exam_id = "jmchc-" + found.group(1)
        rnd = read_round(path)
        if not rnd:
            continue
        if exam_id not in best or len(rnd.scores) > len(best[exam_id].scores):
            best[exam_id] = rnd

    exams, warn = {}, []
    for exam_id, rnd in best.items():
        hist = collections.Counter(rnd.scores)
        n = len(rnd.scores)
        # 문항별 정답자 수의 합 = 전원의 맞은 문항 수 합. 어긋나면 열을 잘못 읽은 것이다.
        if sum(rnd.qc) != sum(rnd.scores):
            warn.append(f"{exam_id}: 정답자 합 {sum(rnd.qc)} ≠ 총점 합 {sum(rnd.scores)}")
        for i, o in enumerate(rnd.qopt):
            lost = rnd.dirty.get(i + 1, 0)
            if o is not None and sum(o) + lost != n:
                warn.append(f"{exam_id} {i+1}번: 선택 분포 합 {sum(o)}+덮임 {lost} ≠ 응시 {n}")
        warn += cross_check(exam_id, rnd)

        exams[exam_id] = {
            "n": n,
            "hist": {str(k): hist[k] for k in sorted(hist)},
            "qc": rnd.qc,
            "q": rnd.qopt,
        }
        gone = [q for q in rnd.dirty if rnd.qopt[q - 1] is None]
        part = [q for q in rnd.dirty if rnd.qopt[q - 1] is not None]
        lost = ((f"  · 분포 없음 {gone}" if gone else "")
                + (f"  · 일부 덮임 {part}" if part else ""))
        print(f"  {exam_id:12s} {n:3d}명  평균 {sum(rnd.scores)/n:4.1f}"
              f"  범위 {min(rnd.scores)}~{max(rnd.scores)}   ← {rnd.path.name}{lost}")

    if warn:
        print("\n  ! 확인이 필요하다:")
        for w in warn:
            print("    " + w)

    return {
        "note": ("회차별 익명 집계. n=응시 인원 · hist=맞은 문항 수별 사람 수 · "
                 "qc=문항별 정답자 수 · q=문항별 [①②③④,무응답] 사람 수. "
                 "이름·학교는 담지 않는다. 엑셀이 '모두정답' 등으로 학생 답을 덮어쓴 "
                 "칸은 빠지므로 q 의 합이 n 보다 작을 수 있다 — 그 문항의 분포 분모는 "
                 "n 이 아니라 q 의 합이다. 한 칸도 안 남은 문항만 null 이다."),
        "source": "조준모의고사 성적표 엑셀의 '성적입력' 시트",
        "exams": dict(sorted(exams.items(), key=lambda kv: _order(kv[0]))),
    }


def _order(exam_id: str) -> tuple[int, int]:
    nums = re.findall(r"\d+", exam_id)
    return (int(nums[0]) if nums else 0, int(nums[1]) if len(nums) > 1 else 0)


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print(__doc__)
        return 2
    folder = Path(args[0]).expanduser()
    if not folder.is_dir():
        print(f"디렉터리가 아니다: {folder}", file=sys.stderr)
        return 2
    force_id = None
    for i, a in enumerate(sys.argv[1:]):
        if a == "--exam" and i + 2 <= len(sys.argv[1:]):
            force_id = sys.argv[1:][i + 1]
        elif a.startswith("--exam="):
            force_id = a.split("=", 1)[1]
    data = build(folder, force_id)
    total = sum(e["n"] for e in data["exams"].values())
    holes = sum(1 for e in data["exams"].values() for o in e["q"] if o is None)
    cells = sum(len(e["q"]) for e in data["exams"].values())
    print(f"\n{len(data['exams'])}개 회차 · {total}명 · 문항 {cells}개"
          f"(선택 분포 없음 {holes}개)")

    # 문항 배열은 한 줄로 — 사람이 읽을 것도 아니고 줄바꿈이 파일을 20배로 불린다.
    text = json.dumps(data, ensure_ascii=False, indent=1)
    text = re.sub(r"\[\s+((?:-?\d+|null)(?:,\s+(?:-?\d+|null))*)\s+\]",
                  lambda m: "[" + re.sub(r"\s+", "", m.group(1)) + "]", text)
    text += "\n"
    if "--write" in sys.argv[1:]:
        # ── 덮어쓰기 전에: 있던 것을 잃지 않는가 ─────────────────────────
        # 2026-08-03 04:52 에 이 파일이 갱신되면서 **문항별 통계(q·qc)가 열
        # 회차에서 통째로 사라졌다**(jmchc-1~10). 또래 정답률이 그 회차에서
        # 안 나오게 됐고, 원본 엑셀이 저장소에 없어 되살릴 방법이 없었다 —
        # 지난 판으로 되돌리는 수밖에 없었다.
        #
        # 그러니 **덮어쓰기 전에** 본다. 있던 것이 없어지면 멈춘다.
        # (정말 지울 뜻이면 --force 를 붙인다. 손이 한 번 더 가야 한다.)
        # ── 얹기(--merge) ────────────────────────────────────────────
        # 엑셀은 회차마다 따로 온다. 한 파일만 놓고 돌리면 나머지 열다섯
        # 회차가 통째로 빠진다 — 아래 안전장치가 막지만, 막히면 아무것도
        # 못 넣는다. --merge 는 있던 것을 남기고 **이번에 읽은 회차만** 얹는다.
        if "--merge" in sys.argv[1:] and OUT.exists():
            try:
                keep = json.loads(OUT.read_text(encoding="utf-8"))
            except Exception:
                keep = {"exams": {}}
            merged = dict(keep.get("exams") or {})
            for eid, val in data["exams"].items():
                if eid in merged:
                    print(f"  얹음(덮어씀) {eid}: {merged[eid].get('n')}명 → {val['n']}명")
                else:
                    print(f"  얹음(새로) {eid}: {val['n']}명")
                merged[eid] = val
            data = dict(data)
            data["exams"] = dict(sorted(merged.items()))
            text = json.dumps(data, ensure_ascii=False, indent=1)
            text = re.sub(r"\[\s+((?:-?\d+|null)(?:,\s+(?:-?\d+|null))*)\s+\]",
                          lambda m: "[" + re.sub(r"\s+", "", m.group(1)) + "]", text)
            text += "\n"

        lost = []
        if OUT.exists():
            try:
                prev = json.loads(OUT.read_text(encoding="utf-8")).get("exams", {})
            except Exception:
                prev = {}
            for eid, old in prev.items():
                new = data["exams"].get(eid)
                if new is None:
                    lost.append(f"{eid}: 회차가 통째로 빠진다")
                    continue
                for key in ("q", "qc"):
                    if old.get(key) and not new.get(key):
                        lost.append(f"{eid}: 문항별 통계 {key} 가 사라진다")
                if (old.get("n") or 0) > (new.get("n") or 0):
                    lost.append(f"{eid}: 인원이 {old.get('n')} → {new.get('n')} 로 준다")
        if lost and "--force" not in sys.argv[1:]:
            print("\n덮어쓰지 않았다 — 있던 것이 없어진다:", file=sys.stderr)
            for line in lost[:12]:
                print("   · " + line, file=sys.stderr)
            if len(lost) > 12:
                print(f"   … 외 {len(lost)-12}건", file=sys.stderr)
            print("\n엑셀 원본이 이번에 덜 담겼을 수 있다. 확인하고, 정말 지울 "
                  "뜻이면 --force 를 붙인다.", file=sys.stderr)
            return 1
        OUT.parent.mkdir(parents=True, exist_ok=True)
        OUT.write_text(text, encoding="utf-8")
        print(f"{OUT.relative_to(ROOT)} 에 적었다 ({len(text)/1024:.1f}KB)")
        return 0
    print("\n--write 를 붙이면 파일에 적는다.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
