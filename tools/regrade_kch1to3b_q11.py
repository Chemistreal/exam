#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
kch1to3-b Q11 재채점 스크립트 (2026-07-07)
================================================
감사 결과: kch1to3-b 시험의 11번 채점키가 4(4)에서 1(1)로 정정됨(3중 대조 + 검산 확증).
index.html(라이브)에는 이미 반영됨. 이 스크립트는 '구글 시트에 이미 저장된 과거 응시 기록'을
정정된 키로 다시 채점하기 위한 도구이다.

무엇을 바꾸나 (Q11 한 문항의 키 4 -> 1):
  - Q11 답을 1(1)로 낸 응시자: 예전엔 오답 처리 -> 이제 정답  => 맞은개수 +1, 원점수 +3점
  - Q11 답을 4(4)로 낸 응시자: 예전엔 정답 처리 -> 이제 오답  => 맞은개수 -1, 원점수 -3점
  - 그 외(2,3 또는 미응답 0): 변동 없음
  * 문항 배점 PT=3, 만점(문항수 60)은 불변.
  * 백점환산 = 원점수 / 만점 * 100 으로 재계산.
  * 백분위/석차/전체누적인원은 '코호트 전체 재순위'가 필요하므로 이 스크립트는 손대지 않고,
    영향 받은 행을 표시만 한다(코호트가 바뀌면 상담 시 재순위 안내 권장).

사용법:
  # 1) 먼저 분석만 (아무것도 바꾸지 않음)
  python regrade_kch1to3b_q11.py sheet_export.csv --dry-run

  # 2) 정정본 파일 생성 (원본은 보존, *_regraded.csv 로 저장)
  python regrade_kch1to3b_q11.py sheet_export.csv

입력 파일:
  - 구글 시트를 CSV 또는 TSV로 내보낸 파일. (xlsx도 지원: openpyxl 있으면)
  - 헤더 행은 백엔드 HEADER와 동일:
    시험, 학생이름, 공유링크, 저장시각, 수험번호, 응시일, 학교, 학년,
    원점수, 만점, 백점환산, 백분위, 석차, 전체누적인원, 맞은개수, 영역별 득점, 답안(60)
  - '시험' 열 값이 kch1to3-b를 가리키는 행만 대상. 기본 매칭:
    정확히 "화학1 1-3단원 모의고사 (동형)" 또는 'kch1to3-b'/'1-3'+'동형' 포함.
    (동형 아닌 kch1to3 '화학1 1-3단원 모의고사'는 제외 - 문구로 구분)
"""
import sys, csv, io, re, os, argparse

EXAM_TITLE = "화학1 1-3단원 모의고사 (동형)"   # 시트 '시험' 열에 저장되는 값
Q_INDEX    = 11        # 정정 문항 번호(1-based)
OLD_KEY    = 4         # 기존(오류) 키
NEW_KEY    = 1         # 정정 키
PT         = 3         # 문항 배점
NQ         = 60        # 문항 수(만점 분모)

COL = {  # 헤더명 -> 의미
 'exam':'시험','name':'학생이름','raw':'원점수','max':'만점',
 'pct100':'백점환산','percentile':'백분위','rank':'석차','n':'전체누적인원',
 'correct':'맞은개수','answers':'답안(60)'
}

def rank_pct(value, arr):
    """앱의 rankPct와 동일: 높은 점수 = 높은 백분위, rank 1 = 최고점, 동점 0.5 가중."""
    n=len(arr); below=sum(1 for v in arr if v<value); equal=sum(1 for v in arr if v==value)
    greater=n-below-equal
    pct=((below+(0.5*equal if equal else 0))/n)*100 if n else 0
    return {'rank':greater+1, 'pct':pct, 'n':n}

def _dash_norm(s):
    # 대시류(en/em/horizontal bar)를 ASCII 하이픈으로 정규화. 소스에 리터럴 대시류를 두지 않으려고 코드로 생성.
    for cp in (0x2013, 0x2014, 0x2015):
        s=s.replace(chr(cp), '-')
    return s

def is_target_exam(v):
    v=(v or '').strip()
    if v==EXAM_TITLE: return True
    # 보조 매칭: '동형'이면서 1-3단원(제목이 en-dash 를 써도 정규화로 매칭), 또는 id 문자열
    if 'kch1to3-b' in v: return True
    if ('1-3' in _dash_norm(v)) and ('동형' in v): return True
    return False

def q11_of(ans):
    """답안 문자열에서 Q11(11번째 글자) 반환. 앞에 붙은 ' 제거, 공백 제거."""
    s=(ans or '').strip().lstrip("'").strip()
    s=re.sub(r'[^0-4]','',s)  # 숫자만
    if len(s)>=Q_INDEX: return s[Q_INDEX-1]
    return None

def load_rows(path):
    ext=os.path.splitext(path)[1].lower()
    if ext in ('.xlsx','.xlsm'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("xlsx를 읽으려면 openpyxl이 필요합니다: pip install openpyxl --break-system-packages")
        wb=load_workbook(path, data_only=True); ws=wb.active
        rows=[[('' if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
        return rows, 'xlsx'
    # csv/tsv 자동 감지
    data=open(path, encoding='utf-8-sig').read()
    delim='\t' if (data.count('\t')>data.count(',')) else ','
    rows=list(csv.reader(io.StringIO(data), delimiter=delim))
    return rows, ('tsv' if delim=='\t' else 'csv')

def main():
    ap=argparse.ArgumentParser(description="kch1to3-b Q11 재채점(4->1)")
    ap.add_argument('input', help="구글 시트 export (csv/tsv/xlsx)")
    ap.add_argument('--dry-run', action='store_true', help="분석만, 파일 미생성")
    ap.add_argument('--rerank', action='store_true',
                    help="정정 후 kch1to3-b 코호트 전체의 백분위·석차·전체누적인원까지 재계산(앱 rankPct 동일)")
    ap.add_argument('-o','--out', help="정정본 출력 경로(기본: <입력>_regraded.<ext>)")
    args=ap.parse_args()

    rows, kind = load_rows(args.input)
    if not rows: sys.exit("빈 파일입니다.")
    header=rows[0]
    idx={}
    for key,name in COL.items():
        if name not in header:
            sys.exit(f"헤더에 '{name}' 열이 없습니다. 시트 원본 헤더 그대로 내보냈는지 확인하세요.\n  발견된 헤더: {header}")
        idx[key]=header.index(name)

    up=down=same=skip_other=0
    changed_rows=[]
    target_rows=[]   # kch1to3-b 전체 대상 행(재순위용) - (row참조, 정정후원점수)
    body=rows[1:]
    for r in body:
        if len(r)<=max(idx.values()):
            r=r+['']*(max(idx.values())+1-len(r))
        if not is_target_exam(r[idx['exam']]):
            skip_other+=1; continue
        q=q11_of(r[idx['answers']])
        try: raw=float(r[idx['raw']]) if str(r[idx['raw']]).strip() not in ('','None') else None
        except ValueError: raw=None
        try: cor=int(float(r[idx['correct']])) if str(r[idx['correct']]).strip() not in ('','None') else None
        except ValueError: cor=None
        try: mx=float(r[idx['max']]) if str(r[idx['max']]).strip() not in ('','None') else NQ*PT
        except ValueError: mx=NQ*PT

        delta=0
        if q==str(NEW_KEY):      # 1: 오답 -> 정답
            delta=+PT; up+=1
        elif q==str(OLD_KEY):    # 4: 정답 -> 오답
            delta=-PT; down+=1
        else:
            same+=1

        # 값 갱신(변동 있을 때만)
        if delta!=0:
            note={'name':r[idx['name']].strip(), 'q11':q, 'delta':delta}
            if raw is not None:
                new_raw=raw+delta
                note['raw_before']=raw; note['raw_after']=new_raw
                r[idx['raw']]=('%g'%new_raw)
                if mx:
                    r[idx['pct100']]=('%.1f'%(new_raw/mx*100))
                    note['pct_after']=round(new_raw/mx*100,1)
            if cor is not None:
                new_cor=cor+(1 if delta>0 else -1)
                note['correct_before']=cor; note['correct_after']=new_cor
                r[idx['correct']]=str(new_cor)
            changed_rows.append(note)

        # 재순위용: 정정 후 원점수(정정 없으면 원래 원점수)로 대상 코호트에 편입
        eff_raw = (raw+delta) if (raw is not None) else None
        target_rows.append((r, eff_raw))

    # ---- 재순위(--rerank): 정정 후 코호트 백분위·석차·전체누적인원 재계산 ----
    reranked=0
    if args.rerank and target_rows:
        totals=[t for (_,t) in target_rows if t is not None]
        N=len(totals)
        for (r,t) in target_rows:
            if t is None: continue
            st=rank_pct(t, totals)
            r[idx['percentile']]=('%.1f'%st['pct'])
            r[idx['rank']]=str(st['rank'])
            r[idx['n']]=str(st['n'])
            reranked+=1

    # 리포트
    print("="*64)
    print("kch1to3-b Q11 재채점 분석  (키 4 -> 1)")
    print("="*64)
    print(f"입력: {args.input}  ({kind}, 데이터 {len(body)}행)")
    print(f"대상 시험 매칭 행: {up+down+same}  (그 외 시험 {skip_other}행 무시)")
    print(f"  가점(+{PT}) 대상 [Q11=1, 예전 오답->정답]: {up}명")
    print(f"  감점(-{PT}) 대상 [Q11=4, 예전 정답->오답]: {down}명")
    print(f"  변동 없음 [Q11=2/3/미응답]:               {same}명")
    if changed_rows:
        print("\n영향 받은 응시자(원점수/맞은개수 변경):")
        for n in changed_rows:
            rb=n.get('raw_before','?'); ra=n.get('raw_after','?')
            cb=n.get('correct_before','?'); ca=n.get('correct_after','?')
            print(f"  - {n['name'] or '(무명)'}: Q11={n['q11']}  원점수 {rb}->{ra} ({'+' if n['delta']>0 else ''}{n['delta']}) · 맞은개수 {cb}->{ca}")
        if args.rerank:
            print(f"\n[재순위] --rerank 적용: kch1to3-b 대상 {reranked}명의 백분위·석차·전체누적인원을 정정 점수 기준으로 재계산했습니다.")
            print("         (앱과 동일한 rankPct: 높은 점수=높은 백분위, 석차 1=최고점, 동점 0.5 가중. 전체누적인원 = 이 파일의 kch1to3-b 응시 수.)")
            print("         ※ 파일이 부분 export면 순위가 해당 부분 기준입니다. 시트 전체를 내보내 적용하세요.")
        else:
            print("\n[주의] 백분위·석차·전체누적인원은 코호트 전체 재순위가 필요해 이 스크립트가 바꾸지 않았습니다.")
            print("       위 인원의 점수가 바뀌면 순위도 소폭 이동합니다. --rerank 를 붙이면 함께 재계산합니다.")
    else:
        print("\n정정으로 점수가 바뀌는 응시자가 없습니다. (Q11에 1 또는 4로 답한 기록 없음)")
        if args.rerank and reranked:
            print(f"[재순위] 점수 변동은 없지만 --rerank로 {reranked}명의 순위를 재확인했습니다(값 동일할 수 있음).")

    if args.dry_run:
        print("\n--dry-run: 파일을 생성하지 않았습니다.")
        return
    if not changed_rows and not (args.rerank and reranked):
        print("\n변경 사항이 없어 정정본을 만들지 않았습니다.")
        return

    out=args.out
    if not out:
        base,ext=os.path.splitext(args.input); out=base+"_regraded"+(ext if ext else ".csv")
    if kind=='xlsx':
        try:
            from openpyxl import Workbook
        except ImportError:
            sys.exit("xlsx 출력에 openpyxl 필요")
        wb=Workbook(); ws=wb.active
        for r in [header]+body: ws.append(r)
        if not out.lower().endswith(('.xlsx','.xlsm')): out=os.path.splitext(out)[0]+".xlsx"
        wb.save(out)
    else:
        delim='\t' if kind=='tsv' else ','
        with open(out,'w',encoding='utf-8-sig',newline='') as f:
            w=csv.writer(f, delimiter=delim); w.writerow(header)
            for r in body: w.writerow(r)
    print(f"\n정정본 저장: {out}")
    print("  (원본은 그대로 보존됩니다. 시트에 반영하려면 정정본의 해당 열을 확인 후 붙여넣기 하세요.)")

if __name__=='__main__':
    main()
