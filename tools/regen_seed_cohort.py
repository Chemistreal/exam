# -*- coding: utf-8 -*-
"""동형 2종(kch1to3-b, kch1to2-b) 시드 코호트 영역벡터 재생성.

2026-07 분류 검수에서 영역 재배정 4건이 라이브에 반영됨:
  kch1to3-b #3  기타 -> 화학의 기초
  kch1to3-b #19 양적관계 -> 기체
  kch1to3-b #52 분자의 극성 -> 분자의 모양
  kch1to2-b #2  원자 -> 보어모형
시드 매트릭스는 영역별 합계만 저장하므로 재배정을 소급하려면 원시 답안이 필요.
구글 시트 '성적기록' 탭 export(csv/tsv/xlsx, 헤더에 '시험'과 '답안(60)' 포함)를 입력하면
새 매핑으로 영역벡터를 다시 계산해 index.html에 붙여넣을 JSON을 출력한다.

채점 키와 영역 매핑은 하드코딩하지 않고 index.html(라이브)에서 직접 읽는다.
따라서 이후 분류가 또 바뀌어도 이 스크립트는 수정 없이 재사용 가능.

사용법:
    python3 regen_seed_cohort.py sheet_export.csv                     # 분석 + JSON 출력
    python3 regen_seed_cohort.py sheet_export.csv --write out.json
    python3 regen_seed_cohort.py sheet_export.csv --index ../index.html   # 경로 지정

참고: 같은 export로 Q11 재채점(regrade_kch1to3b_q11.py)도 함께 돌리는 것을 권장
(둘 다 동일한 원시 답안을 쓰므로 한 번의 내보내기로 충분).
"""
import sys, os, csv, json, io, re

EXAM_KEYS = {
    "kch1to3-b": {"qvar": "Q_KCH13B", "match": lambda t: "1-3" in t and "동형" in t},
    "kch1to2-b": {"qvar": "Q_KCH12B", "match": lambda t: ("1-2" in t or "1,2" in t) and "동형" in t},
}

def _extract_array(html, name):
    i = html.find(name)
    if i < 0:
        sys.exit("index.html에서 %s 를 찾지 못했습니다." % name)
    i = html.find("[", i)
    depth = 0
    for j in range(i, len(html)):
        if html[j] == "[":
            depth += 1
        elif html[j] == "]":
            depth -= 1
            if depth == 0:
                return json.loads(html[i:j+1])
    sys.exit("%s 배열 파싱 실패" % name)

def load_live(index_path):
    html = open(index_path, encoding="utf-8").read()
    live = {}
    for eid, cfg in EXAM_KEYS.items():
        rows = _extract_array(html, cfg["qvar"])
        if len(rows) != 60:
            sys.exit("%s 문항 수 이상: %d" % (cfg["qvar"], len(rows)))
        ans = {r[0]: r[1] for r in rows}
        area = {r[0]: r[3] for r in rows}
        m = re.search(r'id:"%s".*?areaOrder\s*:\s*(\[[^\]]*\])' % re.escape(eid), html, re.DOTALL)
        if not m:
            sys.exit("%s areaOrder 추출 실패" % eid)
        order = json.loads(m.group(1))
        orphan = set(area.values()) - set(order)
        if orphan:
            sys.exit("%s 고아 영역 %s (index.html 정합 확인 필요)" % (eid, orphan))
        live[eid] = {"ans": ans, "area": area, "order": order, "match": cfg["match"]}
    return live

def load_rows(path):
    if path.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("xlsx 입력에는 openpyxl이 필요합니다. csv/tsv로 내보내 주세요.")
        ws = load_workbook(path, read_only=True).active
        rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
    else:
        raw = open(path, encoding="utf-8-sig").read()
        delim = "\t" if raw.count("\t") > raw.count(",") else ","
        rows = list(csv.reader(io.StringIO(raw), delimiter=delim))
    return rows[0], rows[1:]

def main():
    args = sys.argv[1:]
    if not args:
        sys.exit(__doc__)
    path = args[0]
    out_path = args[args.index("--write")+1] if "--write" in args else None
    idx_path = args[args.index("--index")+1] if "--index" in args else \
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "index.html")
    live = load_live(idx_path)

    header, body = load_rows(path)
    try:
        i_exam = header.index("시험")
        i_ans = next(i for i, hh in enumerate(header) if hh.startswith("답안"))
    except (ValueError, StopIteration):
        sys.exit("헤더에 '시험'과 '답안(60)' 열이 필요합니다. 현재 헤더: %s" % header)

    result = {}
    for eid, cfg in live.items():
        rows = [r for r in body if len(r) > max(i_exam, i_ans) and cfg["match"](r[i_exam])]
        mat, skipped = [], 0
        for r in rows:
            a = (r[i_ans] or "").strip()
            if len(a) < 60:
                skipped += 1
                continue
            vec = {ar: 0 for ar in cfg["order"]}
            for q in range(1, 61):
                ch = a[q-1]
                if ch.isdigit() and int(ch) == cfg["ans"][q]:
                    vec[cfg["area"][q]] += 3
            mat.append([vec[ar] for ar in cfg["order"]])
        result[eid] = mat
        print("%s: 대상 %d행, 답안 불량 제외 %d, 벡터 %d개 생성" % (eid, len(rows), skipped, len(mat)))
        print("  index.html의 기존 시드 행수와 대조해 인원 수가 맞는지 확인하세요.")

    payload = ",\n".join('"%s":%s' % (eid, json.dumps(m, separators=(",", ":"))) for eid, m in result.items())
    if out_path:
        open(out_path, "w", encoding="utf-8").write(payload)
        print("저장:", out_path)
    else:
        print("\n----- index.html 시드 교체용 JSON (해당 키 부분만 바꿔 넣기) -----")
        print(payload[:2000] + ("\n... (--write로 전체 저장)" if len(payload) > 2000 else ""))

if __name__ == "__main__":
    main()
